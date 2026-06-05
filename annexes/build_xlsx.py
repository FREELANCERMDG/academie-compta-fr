# -*- coding: utf-8 -*-
"""Genere le classeur Excel 'Outils-Formation.xlsx' (10 onglets) avec en-tetes,
formules, listes deroulantes et mise en forme. Necessite openpyxl."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

OUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(OUT_DIR, "excel", "Outils-Formation.xlsx")
os.makedirs(os.path.dirname(OUT), exist_ok=True)

# ---- Styles ----
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUB_FONT = Font(italic=True, size=9, color="555555")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

STATUTS = '"A FAIRE,EN COURS,BLOQUE,A VALIDER,TERMINE"'
OUINON = '"O,N"'

wb = Workbook()

def setup(ws, title, subtitle, headers, widths, header_row=3):
    ws["A1"] = title; ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle; ws["A2"].font = SUB_FONT
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.fill = HEAD_FILL; c.font = HEAD_FONT; c.alignment = CENTER; c.border = BORDER
        ws.column_dimensions[chr(64+i) if i <= 26 else 'A'+chr(64+i-26)].width = w
    ws.freeze_panes = ws.cell(row=header_row+1, column=1)
    return header_row + 1

def fmt_rows(ws, first_data_row, ncols, nrows=40):
    for r in range(first_data_row, first_data_row + nrows):
        for c in range(1, ncols+1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER; cell.alignment = LEFT

# 1) Suivi dossiers
ws = wb.active; ws.title = "Suivi dossiers"
h = ["Client","Mois","Pieces recues","Releves complets","Saisie achats","Saisie ventes","Banque","Rapprochement","Lettrage","TVA preparee","TVA validee","Pieces manquantes","Anomalies","Statut","Date livraison"]
w = [18,10,12,14,12,12,10,13,10,12,11,18,18,12,13]
fr = setup(ws,"SUIVI DE DOSSIER MENSUEL","Statuts: A FAIRE / EN COURS / BLOQUE / A VALIDER / TERMINE",h,w)
fmt_rows(ws,fr,len(h))
dv = DataValidation(type="list", formula1=STATUTS, allow_blank=True); ws.add_data_validation(dv); dv.add(f"N{fr}:N{fr+39}")
dv2 = DataValidation(type="list", formula1=OUINON, allow_blank=True); ws.add_data_validation(dv2)
for col in ["C","D"]: dv2.add(f"{col}{fr}:{col}{fr+39}")
ws.conditional_formatting.add(f"N{fr}:N{fr+39}", CellIsRule(operator="equal", formula=['"TERMINE"'], fill=PatternFill("solid",fgColor="C6EFCE")))
ws.conditional_formatting.add(f"N{fr}:N{fr+39}", CellIsRule(operator="equal", formula=['"BLOQUE"'], fill=PatternFill("solid",fgColor="FFC7CE")))
ws[f"A{fr}"]="(exemple) CLIENT MARTIN"; ws[f"B{fr}"]="2026-03"; ws[f"N{fr}"]="EN COURS"

# 2) Tableau de bord (KPI)
ws = wb.create_sheet("Tableau de bord")
h = ["Client","Echeance TVA","Pieces recues le","Saisie","Banque","TVA","Livre le","Valide","Retour (O/N)","Anomalies","Temps passe (h)"]
w = [18,14,15,10,10,8,12,9,12,18,14]
fr = setup(ws,"TABLEAU DE BORD DE PRODUCTION (KPI)","Cible taux de retour < 5%",h,w)
fmt_rows(ws,fr,len(h))
dvr = DataValidation(type="list", formula1=OUINON, allow_blank=True); ws.add_data_validation(dvr); dvr.add(f"I{fr}:I{fr+39}")
last = fr+39
ws[f"A{last+2}"]="INDICATEURS"; ws[f"A{last+2}"].font=Font(bold=True,color="1F4E78")
ws[f"A{last+3}"]="Dossiers livres"; ws[f"B{last+3}"]=f'=COUNTA(G{fr}:G{last})'
ws[f"A{last+4}"]="Dossiers avec retour"; ws[f"B{last+4}"]=f'=COUNTIF(I{fr}:I{last},"O")'
ws[f"A{last+5}"]="Taux de retour"; ws[f"B{last+5}"]=f'=IF(B{last+3}=0,0,B{last+4}/B{last+3})'; ws[f"B{last+5}"].number_format="0.0%"
ws[f"A{last+6}"]="Total heures"; ws[f"B{last+6}"]=f'=SUM(K{fr}:K{last})'

# 3) Rapprochement bancaire
ws = wb.create_sheet("Rapprochement")
ws["A1"]="ETAT DE RAPPROCHEMENT BANCAIRE"; ws["A1"].font=TITLE_FONT
rows = [
 ("", "Cote COMPTA (512)", "Cote BANQUE (releve)"),
 ("Solde de depart", 0, 0),
 ("(+) Encaissements non credites", 0, 0),
 ("(-) Cheques emis non debites", 0, 0),
 ("(+) Operations sur releve non saisies", 0, 0),
]
r0=3
for i,(a,b,c) in enumerate(rows):
    rr=r0+i
    ws.cell(rr,1,a); ws.cell(rr,2,b); ws.cell(rr,3,c)
    if i==0:
        for col in (1,2,3):
            cc=ws.cell(rr,col); cc.fill=HEAD_FILL; cc.font=HEAD_FONT; cc.alignment=CENTER
# Soldes corriges : compta = depart + encaiss non credites - cheques ... (cote compta deja a jour -> on demontre l'egalite)
res=r0+len(rows)
ws.cell(res,1,"SOLDE CORRIGE").font=Font(bold=True)
ws.cell(res,2,f'=B{r0+1}')  # solde compta corrige (deja a jour)
ws.cell(res,3,f'=C{r0+1}+C{r0+2}-C{r0+3}+C{r0+4}')
ws.cell(res+1,1,"CONTROLE").font=Font(bold=True)
ws.cell(res+1,2,f'=IF(ROUND(B{res}-C{res},2)=0,"OK : equilibre","ECART A CHERCHER")')
ws.column_dimensions["A"].width=38; ws.column_dimensions["B"].width=20; ws.column_dimensions["C"].width=22
ws["A2"]="Renseigner les soldes ; le controle indique si le rapprochement est equilibre."; ws["A2"].font=SUB_FONT

# 4) Balance agee
ws = wb.create_sheet("Balance agee")
h=["Tiers","Total du","0-30 j","31-60 j","61-90 j","> 90 j","Statut"]
w=[24,14,12,12,12,12,16]
fr=setup(ws,"BALANCE AGEE (clients / fournisseurs)","Relancer si > 60 j",h,w)
fmt_rows(ws,fr,len(h))
for r in range(fr,fr+40):
    ws.cell(r,2).value=f'=SUM(C{r}:F{r})'
    ws.cell(r,7).value=f'=IF(F{r}>0,"A RELANCER (>90j)",IF(E{r}>0,"A SURVEILLER",""))'
ws.conditional_formatting.add(f"F{fr}:F{fr+39}", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid",fgColor="FFC7CE")))

# 5) Pieces manquantes
ws = wb.create_sheet("Pieces manquantes")
h=["Client","Date detection","Operation (montant/tiers/date)","Type de piece","Demande le","Relance le","Recu (O/N)","Statut"]
w=[18,14,32,16,13,13,12,14]
fr=setup(ws,"SUIVI DES PIECES MANQUANTES","",h,w)
fmt_rows(ws,fr,len(h))
dvp=DataValidation(type="list",formula1=OUINON,allow_blank=True); ws.add_data_validation(dvp); dvp.add(f"G{fr}:G{fr+39}")

# 6) Immobilisations
ws = wb.create_sheet("Immobilisations")
h=["Bien","Date mise en service","Base HT","Duree (ans)","Taux","Dotation N","Amort. cumules","VNC"]
w=[26,18,14,12,10,14,16,14]
fr=setup(ws,"FICHIER DES IMMOBILISATIONS","Dotation = Base / Duree ; VNC = Base - Amort. cumules",h,w)
fmt_rows(ws,fr,len(h))
for r in range(fr,fr+40):
    ws.cell(r,5).value=f'=IF(D{r}=0,"",1/D{r})'; ws.cell(r,5).number_format="0.0%"
    ws.cell(r,6).value=f'=IF(D{r}=0,"",C{r}/D{r})'
    ws.cell(r,8).value=f'=IF(C{r}="","",C{r}-G{r})'
ws[f"A{fr+41}"]="TOTAUX"; ws[f"A{fr+41}"].font=Font(bold=True)
ws[f"C{fr+41}"]=f'=SUM(C{fr}:C{fr+39})'; ws[f"G{fr+41}"]=f'=SUM(G{fr}:G{fr+39})'; ws[f"H{fr+41}"]=f'=SUM(H{fr}:H{fr+39})'

# 7) Suivi 471
ws = wb.create_sheet("Suivi 471")
h=["Date","Montant","Sens (D/C)","Libelle banque","Nature presumee","Action","Relance le","Statut"]
w=[12,12,11,28,22,22,12,14]
fr=setup(ws,"SUIVI DES COMPTES D'ATTENTE (471)","Objectif: tout passer a IMPUTE avant cloture",h,w)
fmt_rows(ws,fr,len(h))
dvs=DataValidation(type="list",formula1='"EN ATTENTE,IMPUTE,A VALIDER"',allow_blank=True); ws.add_data_validation(dvs); dvs.add(f"H{fr}:H{fr+39}")

# 8) Controle paie
ws = wb.create_sheet("Controle paie")
h=["Mois","Brut (OD)","Brut (bulletins)","Brut (DSN)","Net (OD)","Net verse (banque)","Cotis. (OD)","Cotis. (DSN)","Cotis. payees","Ecart","OK ?"]
w=[10,12,14,12,12,16,12,12,14,10,8]
fr=setup(ws,"CONTROLE PAIE (rapprochement 4 sources)","Ecart = Brut OD - Brut DSN",h,w)
fmt_rows(ws,fr,len(h))
for r in range(fr,fr+40):
    ws.cell(r,10).value=f'=B{r}-D{r}'
    ws.cell(r,11).value=f'=IF(ROUND(J{r},2)=0,"OK","VERIFIER")'

# 9) Carnet erreurs
ws = wb.create_sheet("Carnet erreurs")
h=["Date","Dossier","Erreur","Cause","Correction","Action preventive"]
w=[12,18,28,24,28,28]
fr=setup(ws,"CARNET D'ERREURS PERSONNEL","Outil n1 de progression : ne pas reproduire une erreur",h,w)
fmt_rows(ws,fr,len(h))

# 10) Planning multi-clients
ws = wb.create_sheet("Planning")
h=["Client","Regime TVA","Echeance","Reception pieces cible","Livraison cible","Charge estimee (h)","Semaine"]
w=[18,16,12,20,16,16,10]
fr=setup(ws,"PLANNING MENSUEL MULTI-CLIENTS","Lisser la charge sur le mois",h,w)
fmt_rows(ws,fr,len(h))

wb.save(OUT)
print("OK ->", OUT)
